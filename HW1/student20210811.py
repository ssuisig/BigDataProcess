#!usr/bin/python3
from openpyxl import load_workbook
wb = load_workbook(filename='student.xlsx')

list1=[]
list2=[]
listGrade=[]
listA1=[]
listB1=[]
listC1=[]
A1=0
B1=0
C1=0
A2=0
B2=0
C2=0
listGrade.append(('0',0))

def scoreCheck(i):
    startIndex = list2.index(listGrade[i][0])+1
    ws.cell(row=list2[startIndex:].index(listGrade[i][0])+1+startIndex, column=8, value=listGrade[i][1])

ws = wb.active
for row in range(2, 76):
    i = 0
    sum = 0
    for col in range(3, 7):
        ratio = [0.3, 0.35, 0.34, 0.01]
        sum += ws.cell(column=col, row=row).value*ratio[i]
        ws.cell(row=row, column=7, value=sum)
        i += 1
    list1.append(sum)
list2 = list1.copy()
list2.insert(0, 0)
list1.sort(reverse=True)
list1.insert(0, 0)

for i in range(1, len(list1)):
    lenA = int((len(list1)-1)*0.3)
    lenB = int((len(list1)-1)*0.7)
    lenC = len(list1)-1
    if list1[i] < 40:
        if(list1[i], 'F') in listGrade:
            listGrade.append((list1[i], 'F'))
            scoreCheck(i)
        else:
            listGrade.append((list1[i], 'F'))
            ws.cell(row=list2.index(listGrade[i][0])+1, column=8, value=listGrade[i][1])
    else:
        if i <= lenA:
            if i <= lenA //2:
                if (list1[i], 'A+') in listGrade:
                    listGrade.append((list1[i], 'A+'))
                    scoreCheck(i)
                else:
                    listGrade.append((list1[i], 'A+'))
                    ws.cell(row=list2.index(listGrade[i][0])+1, column=8, value=listGrade[i][1])
                    listA1.append(list1[i])
                A1 += 1
            else:
                if (list1[i], 'A0') in listGrade:
                    listGrade.append((list1[i], 'A0'))
                    scoreCheck(i)
                else:
                    listGrade.append((list1[i],'A0'))
                    ws.cell(row=list2.index(listGrade[i][0])+1, column=8, value=listGrade[i][1])
                A2 += 1
        elif i <= lenB:
            if i <= (lenB-lenA) //2 + lenA:
                if (list1[i], 'B+') in listGrade:
                    listGrade.append((list1[i], 'B+'))
                    scoreCheck(i)
                else:
                    listGrade.append((list1[i], 'B+'))
                    ws.cell(row=list2.index(listGrade[i][0])+1, column=8, value=listGrade[i][1])
                    listB1.append(list1[i])
                B1 += 1
            else:
                if (list1[i], 'B0') in listGrade:
                    listGrade.append((list1[i], 'B0'))
                    scoreCheck(i)
                else:
                    listGrade.append((list1[i],'B0'))
                    ws.cell(row=list2.index(listGrade[i][0])+1, column=8, value=listGrade[i][1])
                B2 += 1
        else:
            if i <= (lenC-lenB) //2 + lenB:
                if (list1[i], 'C+') in listGrade:
                    listGrade.append((list1[i], 'C+'))
                    scoreCheck(i)
                else:
                    listGrade.append((list1[i], 'C+'))
                    ws.cell(row=list2.index(listGrade[i][0])+1, column=8, value=listGrade[i][1])
                    listC1.append(list1[i])
                C1 += 1
            else:
                if (list1[i], 'C0') in listGrade:
                    listGrade.append((list1[i], 'C0'))
                    scoreCheck(i)
                else:
                    listGrade.append((list1[i],'C0'))
                    ws.cell(row=list2.index(listGrade[i][0])+1, column=8, value=listGrade[i][1])
                C2 += 1
if A1 > A2:
    for i in range((A1+A2)//2):
        ws.cell(row=list2.index(listA1[-1])+1, column=8, value='A0')
        listA1.pop()
elif B1 > B2:
    for i in range((B1+B2)//2):
        ws.cell(row=list2.index(listB1[-1])+1, column=8, value='B0')
        listB1.pop()
elif C1 > C2:
    for i in range((C1+C2)//2):
        ws.cell(row=list2.index(listC1[-1])+1, column=8, value='C0')
        listC1.pop()

wb.save('student.xlsx')
